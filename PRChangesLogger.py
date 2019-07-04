"""
    @f1r3f0x - 08/03/2019
    License: MIT

    Azure DevOps PRChangesLogger:
    Creates a file with all the changes in pull request for specific branches.
"""
# Azure
from azure.devops.connection import Connection
from azure.devops.released import git
from azure.devops import exceptions as AZExceptions
from msrest.authentication import BasicAuthentication
from msrest import exceptions as MSExceptions
from azure.devops.v5_0.git.models import GitPullRequestSearchCriteria

# Other
import openpyxl as xl
from tqdm import tqdm
import colorama as color

# Standard library
from pprint import pprint
from datetime import datetime
from concurrent.futures.thread import ThreadPoolExecutor
from concurrent import futures
from collections import namedtuple


MAX_WORKERS = 16  # This will determine the number of pull requests being processed at a time
DEFAULT_PULL_QUANTITY = 9999


config_fields = ['access_token', 'organization_url', 'repository_name', 'pull_quantity']
Config = namedtuple('Config', config_fields, defaults=(None,) * len(config_fields))


def process_pull_requests(git_client, repo_id, pull, ignore_words, ignore_extensionless_files=True) -> dict:
    """
    Worker function, gets all changes from a pull request.
    
    Args:
        git_client (GitClient): Git client to query stuff from respositories.
        repo_id (str): Id from the repository of the PR, is necessary to make the request.
        pull (GitPullRequest): PR to get the changes from.

    Keyword Args:
        ignore_extensionless_files (bool) (default=True): Ignores files without a dot extension.

    Returns:
        dict with pull request changes by file path of the given PR.
    """
    processed_changes = {}
    commits = git_client.get_pull_request_commits(repo_id, pull.pull_request_id)
    
    for commit in commits:
        # TODO: Rewrite in functional
        ignore_commit = False
        for word in ignore_words:
            if commit.comment.lower().find(word) != -1:
                ignore_commit = True
            break

        if not ignore_commit:
            changes = git_client.get_changes(commit.commit_id, repo_id).changes
            for change in changes:
                file_name = change['item']['path']

                if  not '.' in file_name and ignore_extensionless_files:
                    continue

                counter = processed_changes.get(file_name)
                if not counter:
                    counter = 0
                counter +=1
                processed_changes[file_name] = counter

    return processed_changes


def get_changes(access_token, organization_url, target_repo_name, source_branches, target_branch_name, pull_quantity, ignore_words=[]) -> dict:
    """
    Main function, connects to an Azure DevOps Org and gets all the PR changes from the given branches.
    
    Args:
        access_token (str): Azure DevOps access token, must have the necesary permisions.
        organization_url (str): URL of the org to connect to.
        target_repo_name (str): Repo where the PR was made.
        source_branches (list): Names of the source branches of the PRs.
        target_branch_name (str): Name of the target branch of the PRs.
        pull_quantity (int): Quantity of pulls being queried, is a quirk of the API, if you don't give it a number it will not retreive PRs.

    Returns:
        dict with pull request changes by file path.
    """
    print('\nConnecting to API\n')
    try:
        # Create a connection to the org
        credentials = BasicAuthentication('', access_token)
        connection = Connection(base_url=organization_url, creds=credentials)

        # Get git Client
        # See azure.devops.v5_0.models for models
        #     azure.devops.git.git_client_base for git_client methods
        git_client = connection.clients.get_git_client()

        # Get the repo
        repositories = git_client.get_repositories()

    except MSExceptions.ClientRequestError as err:
        print('Client Request Error:', str(err))
        return None
    except MSExceptions.AuthenticationError as err:
        print('Authentication Error: ', str(err))

    target_repo = None
    for repo in repositories:
        if repo.name == target_repo_name:
            target_repo = repo
    
    if not target_repo:
        print(f'Repository {target_repo_name} not found.')
        return None

    all_changes = {}

    for branch in source_branches:

        # Find commits for the specific branch combination
        search_criteria = GitPullRequestSearchCriteria (
            source_ref_name = f'refs/heads/{branch}',
            target_ref_name = f'refs/heads/{target_branch_name}',
            status = 'Completed'
        )

        pull_requests = git_client.get_pull_requests(target_repo.id, search_criteria, top=9999)


        print(f"Proccesing PR commits for {branch}...")
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            future_prs = { executor.submit(process_pull_requests, git_client, target_repo.id, pull, ignore_words): pull for pull in pull_requests}
            for future in tqdm(futures.as_completed(future_prs), unit=' PRs'):
                data = future.result()
                for change in data.keys():
                    if all_changes.get(change):
                        all_changes[change] = all_changes[change] + data[change]
                    else:
                        all_changes[change] = data[change]
        print()

    return all_changes


def create_workbook(source_branches, target_branch_name, changes, workbook_title='output.xlsx') -> xl.Workbook: 
    """
    Creates and saves an Excel Workbook with the list of changes.
    
    Args:
        source_branches (list): Name of the source branch of the PRs.
        target_branch_name (str): Name of the target branch of the PRs.
        changes (dict): Changes by file path.

    Keyword Args:
        workbook_title (str): File path to final workbook.

    Returns:
        (xl.Workbook) In memory workbook.
    """
    print('Creating Excel file...')
    
    wb = xl.load_workbook(filename='format.xlsx', data_only=True)
    sheet = wb.active

    sheet['A1'] = f'{", ".join(source_branches)}   ->   {target_branch_name}'
    sheet['D1'] = 'Generated by PRChangesLogger.py by @f1r3f0x'

    for i, file_name in enumerate(tqdm(sorted(changes.keys()), unit='rows')):
        row_index = i + 3
        sheet.cell(row=row_index, column=1, value=file_name)
        sheet.cell(row=row_index, column=2, value='-')

    file_name = f'changes_to_{target_branch_name}-{datetime.now().timestamp()}.xlsx'

    print(f'Workbook saved as {file_name}\n')
    wb.save(file_name)

    return wb


def get_config(file_path) -> Config:
    """
    Gets the config file and returns a Config tuple.

    Args:
        file_path (str): Path to config file (json)

    Returns:
        (Config) Config namedtupple.
    """
    try:
        config_file = json.load(open(file_path))
        return Config(
            access_token = config_file['access_token'],
            organization_url = config_file['organization_url'],
            repository_name = config_file['repository_name'],
            pull_quantity = int(config_file['pull_quantity'])

        )
    except FileNotFoundError as err:
        print('Config file not found')
        if input('Do you want to create a new one? (Y/N) ').strip().lower() == 'y':
            config = Config(
                access_token = input('Access Token: '),
                organization_url = input('Organization URL: '),
                repository_name = input('Repository Name: ')
            )
            json.dump({
                'access_token': config.access_token,
                'organization_url': config.organization_url,
                'repository_name': config.repository_name,
                'pull_quantity': DEFAULT_PULL_QUANTITY
            }, open('config.json', 'w'))
            print('done!')
            return config
        else:
            print('Closing...')
            exit(0)
    except json.JSONDecodeError as err:
        print('JSON Decoding Error -', str(err))
    except KeyError as err:
        print('Error in JSON Keys:', str(err))
    except ValueError as err:
        print('Pull quantity must be an Int')
    return None


if __name__ == "__main__":
    import json

    color.init(autoreset=True)
    
    print()
    print('F1r3f0x\'s ' + color.Fore.BLUE + color.Style.BRIGHT + color.Back.WHITE + 'Azure DevOps' + color.Style.RESET_ALL + ' PR changes logger')
    print()

    config = get_config('config-test.json')

    if config:
        source_branches = []

        source_branch_name = input('Source Branch name: ')
        source_branches.append(source_branch_name)

        add_more = input('Add more source branches? (Y/N) ')
        if add_more.lower().strip() == 'y':
            adding_branches = True
            while adding_branches:
                source_branches.append(input('New Source Branch name: '))
                
                print('\nCurrent Branches: ')
                pprint(source_branches)
                print()

                add_more = input('Add more source branches? (Y/N) ')

                if add_more.lower().strip() == 'n':
                    adding_branches = False


        target_branch_name = input('Target Branch name: ')

        # TODO: Rewrite in functional
        ignore_words = input('Ignore words (case insensitive, separated by comma): ')
        if ignore_words != '':
            pre_ignore_words = ignore_words.split(',')
            ignore_words = []
            for word in pre_ignore_words:
                ignore_words.append(word.lower().strip())

            changes = get_changes(config.access_token, config.organization_url, config.repository_name,
                source_branches, target_branch_name, config.pull_quantity, ignore_words=ignore_words)
        else:
            changes = get_changes(config.access_token, config.organization_url, config.repository_name,
                source_branches, target_branch_name, config.pull_quantity)

        if changes:
            workbook = create_workbook(source_branches, target_branch_name, changes)
        else:
            print('No changes found')

    print('by @f1r3f0x - https://github.com/F1r3f0x\n')
    input('press enter to close ...\n\n')